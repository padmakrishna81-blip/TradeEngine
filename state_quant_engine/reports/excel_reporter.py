"""Excel report generator."""
from __future__ import annotations
from typing import Any, List
import io
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from state_quant_engine.database.connection import get_session
from state_quant_engine.repositories.trade_log_repository import TradeLogRepository
from state_quant_engine.services.portfolio_service import PortfolioService


def generate_excel_report(settings: Any, scan_results: List = None, version_id: int = 1) -> bytes:
    """Generate a comprehensive Excel workbook report."""
    wb = Workbook()

    _add_summary_sheet(wb, settings, scan_results)
    _add_scanner_sheet(wb, scan_results)
    _add_trade_history_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _style_header(ws, row: int, cols: int) -> None:
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _add_summary_sheet(wb, settings, scan_results) -> None:
    ws = wb.active
    ws.title = "Portfolio Summary"
    ws.append(["STATE Quant Engine - Portfolio Summary"])
    ws.append([f"Generated: {date.today()}"])
    ws.append([])

    svc = PortfolioService(settings)
    summary = svc.get_summary(scan_results)

    ws.append(["Metric", "Value"])
    _style_header(ws, 4, 2)
    ws.append(["Total Capital", f"₹{summary.total_capital:,.0f}"])
    ws.append(["Allocated", f"₹{summary.allocated:,.0f}"])
    ws.append(["Available", f"₹{summary.available:,.0f}"])
    ws.append(["Current Value", f"₹{summary.current_value:,.0f}"])
    ws.append(["MTM P&L", f"₹{summary.mtm:+,.0f}"])
    ws.append(["MTM %", f"{summary.mtm_pct:+.2f}%"])
    ws.append(["Open Positions", summary.open_positions])
    ws.append(["BUY Signals", summary.buy_signals])
    ws.append(["HOLD Signals", summary.hold_signals])
    ws.append(["EXIT Signals", summary.exit_signals])


def _add_scanner_sheet(wb, scan_results) -> None:
    ws = wb.create_sheet("Scanner Results")
    if not scan_results:
        ws.append(["No scan data available"])
        return
    ws.append(["Rank", "Symbol", "Name", "Type", "Price", "Health %", "Signal", "Profit %", "Chunks"])
    _style_header(ws, 1, 9)
    for r in scan_results:
        ws.append([r.rank, r.symbol, r.name, r.asset_type, r.price,
                   round(r.score_pct, 1), r.recommendation, round(r.current_profit, 2), r.chunks_held])

    sig_colors = {"BUY": "92D050", "HOLD": "00B0F0", "WATCH": "FFFF00", "EXIT": "FF0000"}
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            color = sig_colors.get(str(cell.value), "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _add_trade_history_sheet(wb) -> None:
    ws = wb.create_sheet("Trade History")
    session = get_session()
    try:
        repo = TradeLogRepository(session)
        trades = repo.get_all()
        ws.append(["ID", "Date", "Symbol", "Action", "Price", "Quantity", "Remarks"])
        _style_header(ws, 1, 7)
        for t in trades:
            ws.append([t.id, str(t.date), t.symbol, t.action, t.price, t.quantity, t.remarks or ""])
    finally:
        session.close()
